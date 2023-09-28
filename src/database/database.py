import pymongo
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import logging


class SaveSuivi:
    
    def __init__(self, vendeurs: list, excel: str) -> None:
        self.vendeurs = vendeurs
        self.excel = excel

    def transform_df(self, vendeur):
        wb = load_workbook(self.excel)
        sheet_ranges = wb["AGADIR"]
        my_date = {}

        for i in range(9, sheet_ranges.max_row):
            if sheet_ranges[f"C{i}"].value == vendeur:
                my_date.update(
                    {sheet_ranges[f"D{i}"].value: sheet_ranges[f"E{i}"].value}
                )
                
        return my_date
               

    def send_data(self, vendeur: str, data: dict):
        myclient = pymongo.MongoClient("mongodb://localhost:27017/")
        database = myclient["madec"]
        my_collection = database["suivi_journalier"]
        today = date.today().strftime("%d/%m/%Y")
        data_send = {"vendeur": vendeur, "date": today, "data": data}
        test = my_collection.insert_one(data_send)
        
    def send_all_vendeurs(self):
        for vendeur in self.vendeurs:
           #print(self.transform_df(vendeur))
           self.send_data(vendeur, self.transform_df(vendeur))
        logging("All vendeurs send successfully")

    def find_vendeur(self,vendeur):
        myclient = pymongo.MongoClient("mongodb://localhost:27017/")
        database = myclient["madec"]
        my_collection = database["suivi_journalier"]
        for x in my_collection.find({},{"vendeur" :vendeur}):
            print(x)