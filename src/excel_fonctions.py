import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, GradientFill
import json

class Excel:

    def __init__(self, path):
        self.__day_work = 24
        self.path = path

    def get_day_work(self) -> tuple:
        """
        return tuple as total days of month and day works
        """
        wb = load_workbook(self.path)
        sheet_ranges = wb["AGADIR"]
        # wb.active
        day_work = sheet_ranges["C6"].value
        day_work = day_work.split(" ", 2)
        a: str = day_work[0].replace("/", "")
        b: str = day_work[1]

        print(f"day work is : {int(a), int(b)}")
        with open("days.json", "r") as jsonFile:
            data = json.load(jsonFile)

        data["from_file"] = {"t": b, "d": a}

        with open("days.json", "w") as jsonFile:
            json.dump(data, jsonFile)

        return int(a), int(b.strip())

    def fix_sheet(self):

        wb = load_workbook(self.path)
        sheet_ranges_quali = wb["QUALI NV"]
        sheet_ranges_quali.unmerge_cells("E1:K2")
        sheet_ranges_quali.delete_rows(1, 7)
        sheet_ranges_quali.delete_rows(2, 4)
        sheet_ranges_quali.delete_rows(10, 1)
        sheet_ranges_quali.delete_cols(1, 3)
        sheet_ranges_quali.delete_cols(2, 3)
        sheet_ranges_quali.delete_cols(3, 3)
        sheet_ranges_quali.delete_cols(4, 11)
        sheet_ranges_quali.delete_cols(7, 2)
        sheet_ranges_quali.delete_rows(sheet_ranges_quali.max_row - 1)

        sheet_ranges_quali['A1'] = "Vendeur"
        sheet_ranges_quali['C1'] = "ACM"
        sheet_ranges_quali['F1'] = "LINE"
        sheet_ranges_quali['G1'] = "TSM"
        sheet_ranges_quali['G1'].fill = PatternFill("solid", fgColor="4cbb17")
        sheet_ranges_quali["F1"].fill = PatternFill("solid", fgColor="4cbb17")
        ## AGADIR
        sheet_ranges_quanti = wb["AGADIR"]
        sheet_ranges_quanti.unmerge_cells("A8:A9")
        sheet_ranges_quanti.unmerge_cells("B8:B9")
        sheet_ranges_quanti.unmerge_cells("D8:D9")
        sheet_ranges_quanti.unmerge_cells("F8:J8")
        sheet_ranges_quanti.unmerge_cells("K8:O8")
        sheet_ranges_quanti.delete_cols(1, 2)
        #sheet_ranges_quanti.delete_cols(3, 1)
        #sheet_ranges_quanti.delete_cols(7, 2)
        sheet_ranges_quanti.delete_cols(7, 2)
        sheet_ranges_quanti.delete_cols(8, 2)
        sheet_ranges_quanti.delete_cols(10, 1)
        sheet_ranges_quanti.delete_cols(11, 1)

        sheet_ranges_quanti.delete_rows(1, 8)
        sheet_ranges_quanti.delete_rows(2, 32)
        sheet_ranges_quanti['A1'] = "Vendeur"
        sheet_ranges_quanti['B1'] = "Famille"
        sheet_ranges_quanti['C1'] = "J-1"
        sheet_ranges_quanti['D1'] = "REAL"
        sheet_ranges_quanti['E1'] = "OBJ"
        sheet_ranges_quanti['F1'] = "Percent"
        sheet_ranges_quanti['G1'] = "REAL 2024"
        sheet_ranges_quanti['H1'] = "H 2023"
        sheet_ranges_quanti['I1'] = "H %"
        print("max", sheet_ranges_quanti.max_row)
        for i in range(sheet_ranges_quanti.max_row):
            if sheet_ranges_quanti[f"E{i + 1}"].value == '%':
                sheet_ranges_quanti[f"E{i + 1}"].value = 0
            if sheet_ranges_quanti[f"B{i + 1}"].value == 'SAUCES TACOS':
                sheet_ranges_quanti[f"B{i + 1}"].value = 'SAUCES'
        for i in range(sheet_ranges_quanti.max_row):
            if sheet_ranges_quanti[f"H{i + 1}"].value == '%':
                sheet_ranges_quanti[f"H{i + 1}"].value = 0
        for i in range(sheet_ranges_quanti.max_row):
            if sheet_ranges_quanti[f"D{i + 1}"].value is None:
                sheet_ranges_quanti[f"D{i + 1}"].value = 0

            # if sheet_ranges[f"E{i+1}"]=='%':
            # sheet_ranges[f"E{i+1}"]=0
        wb.save("excel/finale.xlsx")
        print("qualitatif & quantitatif saved")
    
    @staticmethod
    def add_data_to_database(data_list):
       
        wb = load_workbook("excel/database.xlsx")
       
        sheet_ranges_database = wb["j-1"]
        max_column = sheet_ranges_database.max_column+1
        for i, data in enumerate(data_list):
            sheet_ranges_database.cell(row=i+1, column=max_column, value=data)
        wb.save("excel/database.xlsx")
        
